from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import connection
from django.http import HttpResponse
from django.core.management import call_command
from .models import Inje_Cost, Inje_Cost_Bile
from .forms import InjeCostForm, InjeCostBileForm
import io
from openpyxl import Workbook
import openpyxl


# 메인 페이지
def home(request):
    data = Inje_Cost.objects.all()
    return render(request, 'index.html', {'data': data})

# Index -> Excel 이동 페이지
def view_index(request):
    data = Inje_Cost.objects.all()
    return render(request, 'view_index.html', {'data': data})

# 데이터 저장 페이지
def input_view(request):
    if request.method == 'POST':
        inp_form = InjeCostForm(request.POST)
        if inp_form.is_valid():
            # 폼에서 계산된 데이터 가져오기
            inje_cost = inp_form.save(commit=False)
            cleaned_data = inp_form.cleaned_data
            
            # 원자재 단가 계산
            def V_Cost_action(texture, grade):
                if texture and grade:
                    if texture == 'ABS':
                        if grade == 'HI121':
                            return 2400
                        else:
                            return 2600
                    elif texture == 'PP': 
                        if grade == 'J-550A':
                            return 1800
                        else:
                            return 1600
                    elif texture == 'PE':
                        if grade == 'M850':
                            return 1700
                        else:
                            return 2100
                    else:
                        return 2300
                else:
                    return 0
            
            # 설비 임률 계산
            def V_Daily_Pro(bright_tone):
                rate_list = {
                    100: 160000,
                    120: 180000,
                    130: 200000,
                    150: 250000,
                    160: 250000,
                    170: 250000,
                    180: 270000,
                    200: 300000,
                    220: 300000,
                    240: 320000,
                    280: 350000,
                    380: 400000,
                    400: 450000
                }
                return rate_list.get(bright_tone, 0)

            # 계산 수행
            product_weight = cleaned_data.get('Product_Weight', 0)
            sr_weight = cleaned_data.get('SR_Weight', 0)
            cav = cleaned_data.get('Cav', 0)
            bright_tone = cleaned_data.get('Bright_tone', 0)
            c_t = cleaned_data.get('C_T', 0)
            texture = cleaned_data.get('Texture', '')
            grade = cleaned_data.get('Grade', '')

            inje_cost.B_Gross_Weight = round((product_weight + sr_weight) / cav, 2) if cav != 0 else 0
            inje_cost.B_Raw_Material_Price = V_Cost_action(texture, grade)
            inje_cost.B_Raw_Material_Cost = round(inje_cost.B_Raw_Material_Price / 1000 * inje_cost.B_Gross_Weight, 2) if inje_cost.B_Raw_Material_Price != 0 else 0
            inje_cost.B_Daily_Production = round((79200 / c_t) * cav, 2) if c_t != 0 else 0
            inje_cost.B_Facility_Utilization_Rate = V_Daily_Pro(bright_tone) if bright_tone in [100, 120, 130, 150, 160, 170, 180, 200, 220, 240, 280, 380, 400] else 0 
            inje_cost.B_Per_unit_Utilization_Cost = round(inje_cost.B_Facility_Utilization_Rate / inje_cost.B_Daily_Production, 2) if inje_cost.B_Daily_Production != 0 else 0
            inje_cost.B_Per_unit_Cost = round(inje_cost.B_Raw_Material_Cost + inje_cost.B_Per_unit_Utilization_Cost, 2)
            inje_cost.B_General_Management_Cost = round((inje_cost.B_Raw_Material_Cost + inje_cost.B_Per_unit_Utilization_Cost) * 0.1, 2)
            inje_cost.B_Defective_Loss = round((inje_cost.B_Raw_Material_Cost + inje_cost.B_Per_unit_Utilization_Cost + inje_cost.B_General_Management_Cost) * 0.1, 2)
            inje_cost.B_Corporate_Profit = round((inje_cost.B_Raw_Material_Cost + inje_cost.B_Per_unit_Utilization_Cost + inje_cost.B_General_Management_Cost + inje_cost.B_Defective_Loss) * 0.1, 2)
            inje_cost.B_Sum_Cost = round(inje_cost.B_Raw_Material_Cost + inje_cost.B_Per_unit_Utilization_Cost + inje_cost.B_General_Management_Cost + inje_cost.B_Defective_Loss + inje_cost.B_Corporate_Profit, 2)
            
            # 모델 인스턴스 저장
            inje_cost.save()
            
            return redirect('home')
    else:
        inp_form = InjeCostForm()
        
    data = Inje_Cost.objects.all()
    return render(request, 'input_page.html',{'data': data, 'inp_form': inp_form})

# 데이터 삭제 페이지
def delete_view(request):
    if request.method == 'POST':
        form = InjeCostForm(request.POST)
        action = request.POST.get('delete')
        item = request.POST.getlist('item')
        if action == 'data_del':
            if item:
                Inje_Cost.objects.filter(Data_Number__in=item).delete()
                return redirect('home')
    else:
        form = InjeCostForm()
    data = Inje_Cost.objects.all()
    return render(request, 'delete_page.html',{'data': data, 'form': form })

# 데이터 견적 입력 페이지
def detail_view(request):
    data_number = request.GET.get('data_number', '')
    mold_number = request.GET.get('mold_number', '')
    product_name = request.GET.get('product_name', '')
    texture = request.GET.get('texture', '')
    grade = request.GET.get('grade', '')
    product_weight = float(request.GET.get('product_weight', 0))
    sr_weight = float(request.GET.get('sr_weight', 0))
    bright_tone = float(request.GET.get('bright_tone', 0))
    cav = float(request.GET.get('cav', 0))
    c_t = float(request.GET.get('c_t', 0))

    request.session['Data_Number'] = request.GET.get('data_number', '')
    request.session['mold_number'] = request.GET.get('mold_number', '')
    request.session['product_name'] = request.GET.get('product_name', '')
    request.session['texture'] = request.GET.get('texture', '')
    request.session['grade'] = request.GET.get('grade', '')
    request.session['product_weight'] = float(request.GET.get('product_weight', 0))
    request.session['sr_weight'] = float(request.GET.get('sr_weight', 0))
    request.session['bright_tone'] = float(request.GET.get('bright_tone', 0))
    request.session['cav'] = float(request.GET.get('cav', 0))
    request.session['c_t'] = float(request.GET.get('c_t', 0))


    # 원자재 단가 계산
    def V_Cost_action(texture, grade):
        if texture and grade:
            if texture == 'ABS':
                if grade == 'HI121':
                    return 2400
                else:
                    return 2600
            elif texture == 'PP': 
                if grade == 'J-550A':
                    return 1800
                else:
                    return 1600
            elif texture == 'PE':
                if grade == 'M850':
                    return 1700
                else:
                    return 2100
            else:
                return 2300
        else:
            return 0

    # 설비 임률 계산
    def V_Daily_Pro(bright_tone):
        list = {
            100: 160000,
            120: 180000,
            130: 200000,
            150: 250000,
            160: 250000,
            170: 250000,
            180: 270000,
            200: 300000,
            220: 300000,
            240: 320000,
            280: 350000,
            380: 400000,
            400: 450000
        }
        return list.get(bright_tone, 0)

    # 부가 계산
    gross_weight = round((product_weight + sr_weight) / cav, 2) if cav != 0 else 0
    raw_material_price = V_Cost_action(texture, grade)
    raw_material_cost = round(raw_material_price / 1000 * gross_weight, 2) if raw_material_price != 0 else 0
    daily_production = round((79200 / c_t) * cav, 2) if c_t != 0 else 0
    facility_utilization_rate = V_Daily_Pro(bright_tone) if bright_tone in [100, 120, 130, 150, 160, 170, 180, 200, 220, 240, 280, 380, 400] else 0 
    per_unit_utilization_cost = round(facility_utilization_rate / daily_production, 2) if daily_production != 0 else 0
    per_unit_cost = round(raw_material_cost + per_unit_utilization_cost, 2)
    general_management_cost = round((raw_material_cost + per_unit_utilization_cost) * 0.1, 2)
    defective_loss = round((raw_material_cost + per_unit_utilization_cost + general_management_cost) * 0.1, 2)
    corporate_profit = round((raw_material_cost + per_unit_utilization_cost + general_management_cost + defective_loss) * 0.1, 2)
    sum_cost = round(raw_material_cost + per_unit_utilization_cost + general_management_cost + defective_loss + corporate_profit, 2)

    # 데이터 보낼 변수
    context = {
        'data_number': data_number,
        'mold_number': mold_number,
        'product_name': product_name,
        'texture': texture,
        'grade': grade,
        'product_weight': round(product_weight, 2),
        'sr_weight': round(sr_weight, 2),
        'gross_weight': gross_weight,
        'raw_material_price': round(raw_material_price, 2),
        'raw_material_cost': raw_material_cost,
        'bright_tone': round(bright_tone, 2),
        'cav': round(cav, 2),
        'c_t': round(c_t, 2),
        'daily_production': daily_production,
        'facility_utilization_rate': facility_utilization_rate,
        'per_unit_utilization_cost': per_unit_utilization_cost,
        'per_unit_cost': per_unit_cost,
        'general_management_cost': general_management_cost,
        'defective_loss': defective_loss,
        'corporate_profit': corporate_profit,
        'sum_cost': sum_cost,
    }

    # 견적서 데이터 저장
    if request.method == 'POST':
        sess = request.session.get('Data_Number')
        bil_form = InjeCostBileForm(request.POST)
        if bil_form.is_valid():
            bile_data = bil_form.save(commit=False)
            bile_data.Bile_Data_Number = sess 
            bile_data.save()
            return redirect('home')
    else:
        bil_form = InjeCostBileForm()

    return render(request, 'detail_page.html', {'context': context, 'bil_form': bil_form, 'data_number': data_number})

# 데이터 견적 확인 페이지
def excel_view(request):
    data_number = request.GET.get('data_number')
    if data_number:
        
        # 원시 SQL문
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM Inje_Cost WHERE Data_Number = %s", [data_number])
            row = cursor.fetchone()
        
        # 결과를 딕셔너리로 변환
        if row:
            inje_cost_data = {
                'Data_Number': row[0],
                'Mold_Number': row[1],
                'Product_Name': row[2],
                'Texture': row[3],
                'Grade': row[4],
                'Product_Weight': row[5],
                'SR_Weight': row[6],
                'B_Gross_Weight': row[7],
                'B_Raw_Material_Price': row[8],
                'B_Raw_Material_Cost': row[9],
                'Bright_tone': row[10],
                'Cav': row[11],
                'C_T': row[12],
                'B_Daily_Production': row[13],
                'B_Facility_Utilization_Rate': row[14],
                'B_Per_Unit_Utilization_Cost': row[15],
                'B_Per_Unit_Cost': row[16],
                'B_General_Management_Cost': row[17],
                'B_Defective_Loss': row[18],
                'B_Corporate_Profit': row[19],
                'B_Sum_Cost': row[20],
            }

        inje_cost_bile_data = get_object_or_404(Inje_Cost_Bile, Bile_Data_Number = data_number)
        context = { 
            'data': inje_cost_data,
            'user_data': inje_cost_bile_data,
        }
        return render(request, 'excel_page.html', context)
    else:
        return render(request, 'home', {'error': 'No data number provided.'})


def export(request):
    # 세션에서 Data_Number 가져오기
    sess = request.GET.get('data_number')
    if not sess:
        return HttpResponse("No Data_Number found in session", status=400)
    # 데이터베이스에서 데이터 조회
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM Inje_Cost WHERE Data_Number = %s", [sess])
        row = cursor.fetchone()
        
    if not row:
        return HttpResponse("No data found for the provided Data_Number", status=404)
    
    inje_cost_bile_data = get_object_or_404(Inje_Cost_Bile, Bile_Data_Number=sess)
    # 데이터 딕셔너리 생성
    inje_cost_data = {
        'Data_Number': row[0],
        'Mold_Number': row[1],
        'Product_Name': row[2],
        'Texture': row[3],
        'Grade': row[4],
        'Product_Weight': row[5],
        'SR_Weight': row[6],
        'B_Gross_Weight': row[7],
        'B_Raw_Material_Price': row[8],
        'B_Raw_Material_Cost': row[9],
        'Bright_tone': row[10],
        'Cav': row[11],
        'C_T': row[12],
        'B_Daily_Production': row[13],
        'B_Facility_Utilization_Rate': row[14],
        'B_Per_Unit_Utilization_Cost': row[15],
        'B_Per_Unit_Cost': row[16],
        'B_General_Management_Cost': row[17],
        'B_Defective_Loss': row[18],
        'B_Corporate_Profit': row[19],
        'B_Sum_Cost': row[20],
    }

    # Excel 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Inje Cost Data"
    headers_and_rows = [
    (['데이터 번호', '금형번호', '사출설비', '취출방식', '가동시간', '포장방법', '작성일자'], 1),
    (['제품명', '규격', 'CAVITY'], 4),
    (['일일 생산가능 수량', '사용 원자재', '원자재 단가'], 7),
    (['제품 중량', '런너중량', '총제품중량', '원자재 비용'], 10),
    (['계산3', '금액3', '비고'], 13),
    (['계산4', '금액4', '비고'], 16),
    (['계산5', '금액5', '비고'], 19),
    (['합계금액'], 21),
    (['기타내용', '최종 견적가', '업체 결정가'], 24) ]
    data_and_rows = [
    ([
        inje_cost_data['Data_Number'],
        inje_cost_data['Mold_Number'],
        inje_cost_bile_data.Injection_Equipment,
        inje_cost_bile_data.Extraction_Method,
        inje_cost_bile_data.Operating_Time,
        inje_cost_bile_data.Packaging_Method,
        inje_cost_bile_data.Date_of_Preparation
    ], 2),
    ([
        inje_cost_data['Product_Name'],
        inje_cost_data['Grade'],
        inje_cost_data['Cav']
    ], 5),
    ([
        inje_cost_data['B_Daily_Production'],
        inje_cost_bile_data.Raw_Materials_Used,
        inje_cost_data['B_Raw_Material_Price']
    ], 8),
    ([
        inje_cost_data['Product_Weight'],
        inje_cost_data['SR_Weight'],
        inje_cost_data['B_Gross_Weight'],
        inje_cost_data['B_Raw_Material_Cost']
    ], 11),
    ([
        inje_cost_data['B_Per_Unit_Utilization_Cost'],
        inje_cost_data['B_Per_Unit_Cost'],
        inje_cost_bile_data.Remarks_1
    ], 14),
    ([
        inje_cost_data['B_Facility_Utilization_Rate'],
        inje_cost_data['B_General_Management_Cost'],
        inje_cost_bile_data.Remarks_2
    ], 17),
    ([
        inje_cost_data['B_Defective_Loss'],
        inje_cost_data['B_Corporate_Profit'],
        inje_cost_bile_data.Remarks_3
    ], 20),
    ([
        inje_cost_data['B_Sum_Cost']
    ], 22),
    ([
        inje_cost_bile_data.Other,
        inje_cost_bile_data.Final_Estimate,
        inje_cost_bile_data.Supplier_Estimate
    ], 25)
    ]
    # for headers, row in headers_and_rows:
    #     ws.append(headers, row=row)
    #     for data, row in data_and_rows:
    #         ws.append()
    for headers, row in headers_and_rows:
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header)

# 워크시트에 데이터를 추가합니다.
    for data, row in data_and_rows:
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    

    # 메모리 버퍼에 저장
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # HTTP 응답으로 반환
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="사출원가견적서_{sess}.xlsx"'
    return response